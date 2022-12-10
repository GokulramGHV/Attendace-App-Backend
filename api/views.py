from pyexpat import model
from statistics import mode
from django.shortcuts import render
from rest_framework.viewsets import ModelViewSet
from api.serializers import *
from api.models import *
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics, status
from django.http import HttpResponse, HttpResponseNotFound
from rest_framework.decorators import api_view
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import permission_classes

# from rest_framework.authentication import SessionAuthentication
from api.authentication import TokenAuthentication


from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from datetime import datetime

# Create your views here.


# from rest_framework.authentication import TokenAuthentication
# from dj_rest_auth.registration.views import LoginView


# class LoginViewCustom(LoginView):
#     authentication_classes = (TokenAuthentication,)


class CourseViewSet(ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return Course.objects.filter(teacher=self.request.user)

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class StudentViewSet(ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return Student.objects.filter(
            course_enrolled__teacher=self.request.user
        ).order_by("reg_no")

    def get_serializer_class(
        self,
    ):  # refer https://stackoverflow.com/questions/41312558/django-rest-framework-post-nested-objects
        if self.request.method in ["GET"]:
            return StudentReadSerializer
        return StudentSerializer


class CourseStudentView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, pk, format=None):
        students = Student.objects.filter(course_enrolled_id=pk).order_by("reg_no")
        print(len(students))
        serializer = StudentSerializer(students, many=True)
        return Response(serializer.data)


class TimeTableViewSet(ModelViewSet):
    queryset = TimeTable.objects.all()
    serializer_class = TimeTableSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return TimeTable.objects.filter(course__teacher=self.request.user).order_by(
            "start_time"
        )

    def get_serializer_class(
        self,
    ):  # refer https://stackoverflow.com/questions/41312558/django-rest-framework-post-nested-objects
        if self.request.method in ["GET"]:
            return TimeTableReadSerializer
        return TimeTableSerializer


class SessionsViewSet(ModelViewSet):
    queryset = Sessions.objects.all()
    serializer_class = SessionsSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return Sessions.objects.filter(course__teacher=self.request.user)

    def get_serializer_class(
        self,
    ):  # refer https://stackoverflow.com/questions/41312558/django-rest-framework-post-nested-objects
        if self.request.method in ["GET"]:
            return SessionsReadSerializer
        return SessionsSerializer


class CourseSessionsView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, pk, format=None):
        sessions = Sessions.objects.filter(course_id=pk)
        serializer = SessionsReadSerializer(sessions, many=True)
        return Response(serializer.data)


class AttendanceViewSet(ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return Attendance.objects.filter(
            session__course__teacher=self.request.user
        ).order_by("student__reg_no")

    def get_serializer_class(
        self,
    ):
        if self.request.method in ["GET"]:
            return AttendanceReadSerializer
        return AttendanceSerializer


class SessionAttendanceView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request, pk, format=None):
        attendance = Attendance.objects.filter(session_id=pk).order_by(
            "student__reg_no"
        )
        serializer = AttendanceReadSerializer(attendance, many=True)
        return Response(serializer.data)


class BulkAttendanceView(generics.ListCreateAPIView):
    permission_classes = (IsAuthenticated,)
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer

    def create(self, request, course_id, session_id, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, many=True)
        student_data = request.data
        student_ids = []
        for student in student_data:
            student_ids.append(student["student"])
        all_students = Student.objects.filter(course_enrolled=course_id).order_by(
            "reg_no"
        )

        sessions = Sessions.objects.filter(course=course_id)
        total_hours = 0
        for sess in sessions:
            total_hours += sess.block_hours

        print("Total Hours: ", total_hours)
        for student in all_students:
            percent = student.attendance_percentage
            session = Sessions.objects.get(id=session_id)
            attended_hours = (total_hours - session.block_hours) * percent
            if student.id not in student_ids:
                attended_hours += session.block_hours
            print(student.name, attended_hours)
            student.attendance_percentage = attended_hours / total_hours
            student.save()

        serializer.is_valid(raise_exception=True)
        try:
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes((IsAuthenticated,))
def waprfile(request, cid, sid):
    workbook = Workbook()
    worksheet = workbook.active
    students = Student.objects.filter(course_enrolled_id=cid).order_by("reg_no")
    absent_studs = Attendance.objects.filter(session=sid).order_by("student__reg_no")
    stud_list = []
    session_data = Sessions.objects.get(id=sid)
    print(absent_studs)
    ab_lst = []
    for stud in absent_studs:
        ab_lst.append(stud.student_id)

    for student in students:
        if student.id in ab_lst:
            stud_list.append(
                {"name": student.name, "attendance": "A(0)", "reg_no": student.reg_no}
            )
        else:
            stud_list.append(
                {
                    "name": student.name,
                    "attendance": f"P({session_data.block_hours})",
                    "reg_no": student.reg_no,
                }
            )

    print("Students:", students)
    for i in range(len(stud_list)):
        worksheet.cell(row=i + 1, column=1, value=stud_list[i]["reg_no"])
        worksheet.cell(row=i + 1, column=2, value=stud_list[i]["name"])
        worksheet.cell(row=i + 1, column=3, value=stud_list[i]["attendance"])

    # worksheet["A1"] = f"Session {}"
    # worksheet["B1"] = "world!"
    response = HttpResponse(
        save_virtual_workbook(workbook),
        # content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": 'attachment; filename="foo.xlsx"',
        },
    )
    # response["Content-Disposition"] = "attachment; filename=myexport.xlsx"
    return response


@api_view(["GET"])
@permission_classes((IsAuthenticated,))
def all_sessions_attendance(request, cid):
    workbook = Workbook()
    worksheet = workbook.active

    students = Student.objects.filter(course_enrolled_id=cid).order_by("reg_no")
    sessions = Sessions.objects.filter(course=cid)
    no_of_sessions = len(sessions)

    total_hours = 0
    for sess in sessions:
        total_hours += sess.block_hours

    worksheet.cell(row=1, column=1, value="Reg No.")
    worksheet.cell(row=1, column=2, value="Name")
    worksheet.cell(row=1, column=no_of_sessions + 3, value="Attended Hours")
    worksheet.cell(row=1, column=no_of_sessions + 4, value="Total Hours")

    for i in range(1, len(students) + 1):
        worksheet.cell(row=i + 1, column=1, value=students[i - 1].reg_no)
        worksheet.cell(row=i + 1, column=2, value=students[i - 1].name)
        worksheet.cell(row=i + 1, column=no_of_sessions + 4, value=total_hours)

    init_row = 2
    init_col = 3

    attendance_dict = {}
    for student in students:
        attendance_dict[student.id] = 0

    for session in sessions:
        absent_studs = Attendance.objects.filter(session=session.id).order_by(
            "student__reg_no"
        )
        worksheet.cell(
            row=1,
            column=init_col,
            value=datetime.fromisoformat(str(session.session)).strftime(
                "%m-%d-%Y %H:%M:%S"
            ),
        )

        ab_lst = []
        for stud in absent_studs:
            ab_lst.append(stud.student_id)

        for student in students:
            if student.id in ab_lst:
                worksheet.cell(row=init_row, column=init_col, value="A(0)")
            else:
                worksheet.cell(
                    row=init_row,
                    column=init_col,
                    value=f"P({session.block_hours})",
                )
                attendance_dict[student.id] += session.block_hours
            init_row += 1
        init_col += 1
        init_row = 2

    init_row = 2
    for stud_id in attendance_dict:
        worksheet.cell(
            row=init_row,
            column=no_of_sessions + 3,
            value=attendance_dict[stud_id],
        )
        init_row += 1

    response = HttpResponse(
        save_virtual_workbook(workbook),
        headers={
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": 'attachment; filename="foo.xlsx"',
        },
    )
    return response
